import openpyxl
import pytest
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
import xlrd

# Create the Chrome Driver instance
driver = webdriver.Chrome()


# Read Login Credentials from an Excel File


def read_login_data(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    login_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        Username, Password = row
        login_data.append((Username, Password))
    return login_data


# Parametrize Test using Pytest
@pytest.mark.parametrize("Username, Password",
                         read_login_data("C:\\Users\\GangaprasadK\\Desktop\\login_data.xlsx", "Sheet1"))
def test_login(Username, Password):   # Argument inside the function always should be in lowercase
    # Open website
    driver.get("https://practice.expandtesting.com/login")

    # Perform the Login Actions
    Username_input = driver.find_element(By.ID, "username")
    Password_input = driver.find_element(By.ID, "password")
    Login_button = driver.find_element(By.XPATH, "//*[@id='login']/button")

    # Fill the username and the Password which are in xlsx file
    Username_input.send_keys(Username)
    Password_input.send_keys(Password)

    # Submit the Form
    Login_button.click()

    # Check if login is successful or not
    if "Invalid credentials" in driver.page_source:
        raise AssertionError(f"Login failed for username: {Username}, password: {Password}")

    print(driver.find_element(By.XPATH,'//*[@id="flash"]/b').text)


    # Print the paragraph present inside the webpage
    logout_page_para = driver.find_element(By.TAG_NAME, "h1")
    print(logout_page_para.text)
    print(driver.find_element(By.CLASS_NAME, "subheader").text)
    # Logout immediately after the login successfully
    logout_button = driver.find_element(By.LINK_TEXT, "Logout")
    logout_button.click()


if __name__ == "__main__":
    pytest.main(["-v", "test_openxl_login.py"])
